"""MISO DPP-2025-Cycle Definitive Planning Phase study reports -- document-route loader.

WHAT THIS IS (measured 2026-08-17)
----------------------------------
The CartoVista POI heatmap's three headline tables (POIs / TSA / GIQueue) are 403 ProtectedData
(see docs/MISO_DPP2025_ROUTE.md section 2 -- exhausted, do not re-probe). This script works the
MISO-AUTHORED document route instead:

  1. DISCOVERY -- MISO's own website document index, an OPEN Elasticsearch endpoint (anonymous
     POST, HTTP 200, verified with plain curl):
        POST https://www.misoenergy.org/api/find/Optics_Models_Find_RemoteHostedContentItem/_search
     Body is standard ES query DSL. Each hit carries ObjectId (the cdn suffix), FileName,
     ContentType, and publisher metadata incl. Properties.studycycle ("2025 Cycle"),
     studygroup, processstage, and the J-numbers covered by the document.

  2. FILE -- the cdn URL is deterministic from the index hit:
        https://cdn.misoenergy.org/{FileName stem}{ObjectId}.{ext}
     e.g. https://cdn.misoenergy.org/GI-DPP-2025-ALL_SIS_Ph1_FINAL_v1.0_PUBLIC_20260324748615.zip
     -> HTTP 200, application/zip, 284,571 bytes, Last-Modified 2026-04-01. (The Name-with-spaces
     variant 403s; use FileName.)

WHAT THE PUBLIC ZIP CONTAINS -- AND WHAT IT DOES NOT
-----------------------------------------------------
  GI DPP 2025 Cycle 1 SIS Phase 1 Final Report.pdf     7-page summary, "March 24th, 2026", v1.0
  Appendix I - Executive Project and Upgrade Cost Summary/Executive Cost Summary.xlsx
       -> 206 project rows: Project (J#), Fuel Type, ERIS MW, NRIS MW, Service Type,
          Total DPP-2025 Phase 1 Network Upgrade Cost ($). THIS IS WHAT LOADS.

  THE CONSTRAINT-LEVEL TABLES ARE NOT PUBLIC. The report's own words:
     "The full list of constraints identified in the ERIS and NRIS analysis are detailed in
      Appendix C (CEII) - ERIS Results and Appendix D (CEII) - NRIS (Deliverability) Results."
  Appendices A-H are all marked (CEII); only Appendix I ships in the PUBLIC zip. Obtaining CEII
  is out of scope by rule. A blocked appendix recorded verbatim is a SUCCESS, not a failure.

VINTAGE (prove-the-cycle rule): the index hit is publisher-tagged Properties.studycycle =
"2025 Cycle"; the report title is "MISO DPP 2025 Phase 1 Final Report" dated 03/24/2026 v1.0;
the zip's Last-Modified is 2026-04-01. Phase 1 analyzed 351 requests, 58,730.20 MW ERIS /
56,458.40 MW NRIS (report Executive Summary).

SCHEDULE CONTEXT (so a re-runner knows what to expect): per the July 2026 IPWG DPP Study
Schedule (cdn 20260721...769072.pdf, schedule version 7/1/2026), DPP-2025 DPP-1 completed
4/14/2026 and DPP-2 completion moved to 4/28/2027 (DPP-3 7/8/2027). So until ~April 2027 the
Phase 1 report is the ONLY DPP-2025 SIS artifact; run --discover periodically and this script
will list (not auto-load) any newly posted DPP-2025 SIS documents.

BOUNDARIES: anonymous read-only GET/POST of public documents, identifying User-Agent,
>= 1.15 s between requests per host, no accounts, no CEII, ASCII-only console output.

USAGE
-----
    python scripts/pull_miso_dpp2025_study.py --discover          # list DPP-2025 docs in MISO's index, NO writes
    python scripts/pull_miso_dpp2025_study.py --load --dry-run    # download + parse, do not touch BigQuery
    python scripts/pull_miso_dpp2025_study.py --load --smoke      # load capped at 25 rows + registry row
    python scripts/pull_miso_dpp2025_study.py --load              # load Appendix I (~206 rows) + registry row
"""
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import argparse
import io
import json
import re
import time
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone

PROJECT = "energy-platfrom"
DATASET = "indiana_app"          # energy.* is READ-ONLY for this workstream
DS = f"{PROJECT}.{DATASET}"

FIND_URL = ("https://www.misoenergy.org/api/find/"
            "Optics_Models_Find_RemoteHostedContentItem/_search")
CDN = "https://cdn.misoenergy.org"
UA = ("DecennialGroup-DataAudit/1.0 (read-only public planning documents; "
      "contact ahenderson@decennialgroup.com)")
MIN_INTERVAL = 1.15
TIMEOUT = 90
MAX_BYTES = 64 * 1024 * 1024

# The proven Phase 1 target (2026-08-17). --load verifies it is still in the index first.
PH1_OBJECT_ID = 748615
PH1_FILENAME = "GI-DPP-2025-ALL_SIS_Ph1_FINAL_v1.0_PUBLIC_20260324.zip"
OUT_TABLE = "in_miso_dpp2025_ph1_project_costs"

CEII_WALL = ("The full list of constraints identified in the ERIS and NRIS analysis are "
             "detailed in Appendix C (CEII) - ERIS Results and Appendix D (CEII) - "
             "NRIS (Deliverability) Results.")

_last = {}


def _throttle(url):
    host = urllib.parse.urlparse(url).netloc
    dt = time.time() - _last.get(host, 0.0)
    if dt < MIN_INTERVAL:
        time.sleep(MIN_INTERVAL - dt)
    _last[host] = time.time()


def _http(url, data=None, ctype=None):
    _throttle(url)
    req = urllib.request.Request(url, data=data)
    req.add_header("User-Agent", UA)
    if ctype:
        req.add_header("Content-Type", ctype)
    with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
        body = r.read(MAX_BYTES + 1)
        if len(body) > MAX_BYTES:
            raise RuntimeError(f"response exceeds {MAX_BYTES} bytes: {url}")
        return r.status, dict(r.headers), body


def es_search(query_dsl):
    status, _, body = _http(FIND_URL, data=json.dumps(query_dsl).encode("utf-8"),
                            ctype="application/json")
    if status != 200:
        raise RuntimeError(f"find API HTTP {status}")
    return json.loads(body)


def cdn_url(filename, object_id):
    stem, dot, ext = filename.rpartition(".")
    if not dot:
        raise ValueError(f"no extension on FileName: {filename!r}")
    return f"{CDN}/{urllib.parse.quote(stem)}{object_id}.{ext}"


def discover(cycle="2025"):
    """List MISO-hosted DPP-{cycle} documents from the open find index. No writes."""
    dsl = {
        "query": {"query_string": {"query": f"DPP AND {cycle}"}},
        "size": 300,
        "_source": ["ObjectId$$number", "FileName$$string", "Name$$string",
                    "ContentType$$string", "SearchPublishDate$$date",
                    "Properties.studycycle", "Properties.studygroup",
                    "Properties.processstage", "Properties.displaytitle"],
    }
    hits = es_search(dsl).get("hits", {}).get("hits", [])
    pat = re.compile(rf"GI[-_ ]?DPP[-_ ]?{cycle}", re.I)
    docs = []
    for h in hits:
        s = h.get("_source", {})
        fn = s.get("FileName$$string") or s.get("Name$$string") or ""
        if not pat.search(fn):
            continue
        p = s.get("Properties") or {}
        docs.append({
            "object_id": s.get("ObjectId$$number"),
            "file_name": fn,
            "content_type": s.get("ContentType$$string"),
            "published": (s.get("SearchPublishDate$$date") or "")[:10],
            "studycycle": p.get("studycycle"),
            "studygroup": p.get("studygroup"),
            "title": p.get("displaytitle"),
            "url": None,
        })
    for d in docs:
        try:
            d["url"] = cdn_url(d["file_name"], d["object_id"])
        except ValueError:
            d["url"] = f"(no extension; cannot build cdn url for {d['file_name']!r})"
    docs.sort(key=lambda d: d["published"], reverse=True)
    print(f"MISO find index: {len(docs)} GI-DPP-{cycle} documents "
          f"(of {len(hits)} loose hits)")
    for d in docs:
        print(f"  {d['published']}  id={d['object_id']}  cycle={d['studycycle']!r}  "
              f"{d['file_name']}")
        print(f"      -> {d['url']}")
    return docs


def fetch_ph1_zip():
    """Verify the Phase 1 doc is still indexed, then download its PUBLIC zip."""
    dsl = {"query": {"term": {"ObjectId$$number": PH1_OBJECT_ID}}, "size": 1}
    hits = es_search(dsl).get("hits", {}).get("hits", [])
    if hits:
        src = hits[0]["_source"]
        fn = src.get("FileName$$string") or PH1_FILENAME
        cyc = (src.get("Properties") or {}).get("studycycle")
        print(f"index check: ObjectId {PH1_OBJECT_ID} present, FileName={fn!r}, "
              f"publisher studycycle tag={cyc!r}")
        if cyc and cyc != "2025 Cycle":
            raise RuntimeError(f"vintage mismatch: index says {cyc!r}, expected '2025 Cycle'")
    else:
        fn = PH1_FILENAME
        print(f"WARNING: ObjectId {PH1_OBJECT_ID} no longer in index; "
              f"trying last-known FileName")
    url = cdn_url(fn, PH1_OBJECT_ID)
    print(f"GET {url}")
    status, headers, body = _http(url)
    print(f"  HTTP {status}, {len(body):,} bytes, "
          f"Last-Modified={headers.get('Last-Modified')}")
    if status != 200:
        raise RuntimeError(f"cdn HTTP {status}")
    return url, headers, body


def parse_appendix_i(zip_bytes, url, last_modified, cap=None):
    """Rows from 'Appendix I - Executive Project and Upgrade Cost Summary'."""
    import openpyxl
    z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    names = z.namelist()
    print("zip members:")
    for i in z.infolist():
        print(f"  {i.file_size:>10,}  {i.filename}")
    xlsx = [n for n in names if n.lower().endswith(".xlsx") and "cost summary" in n.lower()]
    pdfs = [n for n in names if n.lower().endswith(".pdf")]
    if not xlsx:
        raise RuntimeError(f"no 'Cost Summary' xlsx member; members={names}")
    wb = openpyxl.load_workbook(io.BytesIO(z.read(xlsx[0])), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    def snake(h):
        return re.sub(r"[^a-z0-9]+", "_", h.lower()).strip("_") or "col"
    cols = [snake(h) for h in header]
    colmap = dict(zip(cols, header))
    pulled_at = datetime.now(timezone.utc).isoformat()
    out = []
    for r in rows_iter:
        if r[0] is None:                      # blank / spacer rows
            continue
        rec = {c: (v if not hasattr(v, "isoformat") else v.isoformat())
               for c, v in zip(cols, r)}
        rec.update({
            "_source_url": url,
            "_source_object_id": PH1_OBJECT_ID,
            "_source_zip_member": xlsx[0],
            "_report_pdf_member": pdfs[0] if pdfs else None,
            "_study_vintage_disclosed": ("publisher index tag studycycle='2025 Cycle'; "
                                         "report 'MISO DPP 2025 Phase 1 Final Report' "
                                         "v1.0 dated 03/24/2026"),
            "_source_last_modified": last_modified,
            "_column_name_map": json.dumps(colmap, ensure_ascii=True),
            "_pulled_at": pulled_at,
        })
        out.append(rec)
        if cap and len(out) >= cap:
            print(f"  (smoke cap {cap} reached)")
            break
    print(f"parsed {len(out)} data rows from {xlsx[0]!r}; columns: {cols}")
    return out, colmap


def cmd_load(a):
    url, headers, body = fetch_ph1_zip()
    rows, colmap = parse_appendix_i(body, url, headers.get("Last-Modified"),
                                    cap=25 if a.smoke else None)
    if a.dry_run:
        print("dry-run: first 3 rows:")
        for r in rows[:3]:
            print("  " + json.dumps(r, ensure_ascii=True, default=str)[:300])
        print("dry-run: no BigQuery writes.")
        return
    from google.cloud import bigquery
    client = bigquery.Client(project=PROJECT)
    dest = f"{DS}.{OUT_TABLE}"
    client.load_table_from_json(rows, dest, job_config=bigquery.LoadJobConfig(
        write_disposition="WRITE_TRUNCATE", autodetect=True)).result()
    got = list(client.query(f"SELECT COUNT(*) n FROM `{dest}`").result())[0].n
    print(f"loaded {got:,} rows -> {dest}")
    # _registry row in the SAME run (non-negotiable)
    client.query(f"DELETE FROM `{DS}._registry` WHERE table_name=@t",
                 job_config=bigquery.QueryJobConfig(query_parameters=[
                     bigquery.ScalarQueryParameter("t", "STRING", OUT_TABLE)])).result()
    source = url
    method = ("Static public file on cdn.misoenergy.org, discovered via MISO's open "
              "Elasticsearch find API (anonymous POST "
              "https://www.misoenergy.org/api/find/Optics_Models_Find_RemoteHostedContentItem/_search, "
              "query_string 'DPP AND 2025', hit ObjectId 748615; cdn URL = FileName stem + "
              "ObjectId + ext). Whole-file download, no pagination. "
              "RE-SCRAPE COMMAND: python scripts/pull_miso_dpp2025_study.py --load")
    notes = ("Appendix I 'Executive Project and Upgrade Cost Summary' of the DPP-2025-Cycle "
             "Phase 1 System Impact Study PUBLIC report zip (report PDF 'MISO DPP 2025 "
             "Phase 1 Final Report', March 24th 2026, v1.0 Initial Posting 03/24/2026; zip "
             f"Last-Modified {headers.get('Last-Modified')}). VINTAGE: publisher's index tag "
             "Properties.studycycle='2025 Cycle'. One row per J-number: fuel, ERIS/NRIS MW, "
             "service type, total Phase 1 network upgrade cost USD. Columns: "
             + json.dumps(colmap, ensure_ascii=True) +
             ". EXCLUDED (not in the PUBLIC zip, CEII-designated by MISO): Appendices A-H "
             "incl. Appendix C ERIS Results / Appendix D NRIS Results which hold the "
             "constraint-level tables. Report wall verbatim: '" + CEII_WALL + "' "
             "This table is the project-cost dimension, NOT bus/constraint headroom. "
             "DPP-2025 Phase 2 report not expected before ~2027-04-28 (July 2026 IPWG "
             "schedule v7/1/2026); re-run --discover to see new postings.")
    client.query(
        f"""INSERT `{DS}._registry`
            (table_name, source, method, n_rows, gb_scanned, built_at, notes)
            VALUES (@t,@s,@m,@n,0.0,CURRENT_TIMESTAMP(),@notes)""",
        job_config=bigquery.QueryJobConfig(query_parameters=[
            bigquery.ScalarQueryParameter("t", "STRING", OUT_TABLE),
            bigquery.ScalarQueryParameter("s", "STRING", source),
            bigquery.ScalarQueryParameter("m", "STRING", method),
            bigquery.ScalarQueryParameter("n", "INT64", int(got)),
            bigquery.ScalarQueryParameter("notes", "STRING", notes)])).result()
    print(f"_registry row written for {OUT_TABLE}: n_rows={got:,}")


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--discover", action="store_true",
                    help="list DPP-2025 documents in MISO's index; no writes")
    ap.add_argument("--cycle", default="2025",
                    help="cycle year for --discover (default 2025)")
    ap.add_argument("--load", action="store_true",
                    help="download Phase 1 PUBLIC zip, load Appendix I + registry row")
    ap.add_argument("--smoke", action="store_true", help="cap load at 25 rows")
    ap.add_argument("--dry-run", action="store_true", help="parse only, no BigQuery")
    a = ap.parse_args()
    if a.discover:
        discover(a.cycle)
    elif a.load:
        cmd_load(a)
    else:
        ap.print_help()


if __name__ == "__main__":
    main()
